import os
from datetime import datetime

import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image


os.makedirs("charts", exist_ok=True)
os.makedirs("reports", exist_ok=True)

transactions = pd.read_csv("data/transactions.csv")

total_transactions = len(transactions)
total_amount = transactions["Amount"].sum()
average_amount = transactions["Amount"].mean()
declined_transactions = transactions[transactions["Status"] == "Declined"]

print("========== FinGuard Financial Analytics ==========\n")
print(f"Total Transactions: {total_transactions}")
print(f"Total Transaction Amount: ${total_amount:,.2f}")
print(f"Average Transaction Amount: ${average_amount:,.2f}")
print(f"Declined Transactions: {len(declined_transactions)}")

transactions["Risk Score"] = 0
transactions.loc[transactions["Amount"] >2000, "Risk Score"] += 40
transactions.loc[transactions["Country"] != "USA", "Risk Score"] += 30
transactions.loc[transactions["Status"] == "Declined", "Risk Score"] += 30
transactions.loc[transactions["PaymentMethod"] == "Wire Transfer", "Risk Score"] += 20
transactions.loc[transactions["Category"] == "Crypto", "Risk Score"] += 20


def risk_level(score):
    if score >= 90:
        return "High"
    elif score >= 50:
        return "Medium"
    else:
        return "Low"


def recommended_action(score):
    if score >= 90:
        return "Freeze Transaction & Investigate"
    elif score >= 50:
        return "Manual Review Required"
    else:
        return "Approve Transaction"


def get_reason(row):
    reasons = []

    if row["Amount"] > 2000:
        reasons.append("High Amount")
    if row["Country"] != "USA":
        reasons.append("Foreign Transaction")
    if row["Status"] == "Declined":
        reasons.append("Declined")
    if row["PaymentMethod"] == "Wire Transfer":
        reasons.append("Wire Transfer")
    if row["Category"] == "Crypto":
        reasons.append("Crypto")

    return ", ".join(reasons) if reasons else "Normal Activity"


transactions["Risk Level"] = transactions["Risk Score"].apply(risk_level)
transactions["Reason"] = transactions.apply(get_reason, axis=1)
transactions["Recommended Action"] = transactions["Risk Score"].apply(recommended_action)

risk_report = transactions.sort_values(by="Risk Score", ascending=False)

print("\n=========== Fraud Investigation Report ===========\n")
print(
    risk_report[
        [
            "TransactionID",
            "Merchant",
            "Risk Score",
            "Risk Level",
            "Reason",
            "Recommended Action",
        ]
    ]
    .head(25)
    .to_string(index=False)
)

# PNG Charts
plt.style.use("seaborn-v0_8-whitegrid")

colors = [
    "#1f77b4", "#ff7f0e", "#2ca02c", "#d62728",
    "#9467bd", "#8c564b", "#e377c2", "#7f7f7f",
    "#bcbd22", "#17becf"
]


def make_colorful_bar(data, title, xlabel, ylabel, filename, rotation=0):
    fig, ax = plt.subplots(figsize=(10, 6))
    bars = ax.bar(data.index, data.values, color=colors[:len(data)])

    ax.set_title(title, fontsize=18, fontweight="bold", pad=25)
    ax.set_xlabel(xlabel, fontsize=12, fontweight="bold", labelpad=15)
    ax.set_ylabel(ylabel, fontsize=12, fontweight="bold", labelpad=15)
    ax.tick_params(axis="x", rotation=rotation)
    ax.grid(axis="y", linestyle="--", alpha=0.5)

    for bar in bars:
        height = bar.get_height()
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            height,
            f"{int(height)}",
            ha="center",
            va="bottom",
            fontsize=10,
            fontweight="bold",
        )

    plt.tight_layout(pad=3)
    plt.savefig(f"charts/{filename}", dpi=150, bbox_inches="tight")
    plt.close()


make_colorful_bar(
    transactions["Country"].value_counts(),
    "Transaction Distribution by Country",
    "Country",
    "Number of Transactions",
    "countries.png",
    rotation=45,
)

make_colorful_bar(
    transactions["Merchant"].value_counts().head(10),
    "Top Merchants by Transaction Volume",
    "Merchant",
    "Number of Transactions",
    "top_merchants.png",
    rotation=45,
)

make_colorful_bar(
    transactions["Risk Level"].value_counts(),
    "Transaction Risk Distribution",
    "Risk Level",
    "Number of Transactions",
    "risk_levels.png",
    rotation=0,
)

make_colorful_bar(
    transactions["PaymentMethod"].value_counts(),
    "Transaction Distribution by Payment Method",
    "Payment Method",
    "Number of Transactions",
    "payment_methods.png",
    rotation=45,
)

make_colorful_bar(
    transactions["Category"].value_counts(),
    "Transaction Category Distribution",
    "Category",
    "Number of Transactions",
    "transactions_by_category.png",
    rotation=45,
)

print("\nCharts saved to the charts folder.")

# Excel Export
output_file = "reports/FinGuard_Analytics_Report.xlsx"

report_columns = [
    "TransactionID",
    "Merchant",
    "Amount",
    "Country",
    "PaymentMethod",
    "Status",
    "Risk Score",
    "Risk Level",
    "Reason",
    "Recommended Action",
]

risk_report[report_columns].to_excel(output_file, index=False)

workbook = load_workbook(output_file)
sheet = workbook.active
sheet.title = "Fraud Report"

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)

for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

sheet.freeze_panes = "B2"

table = Table(displayName="FraudReportTable", ref=sheet.dimensions)
style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
table.tableStyleInfo = style
sheet.add_table(table)

sheet.conditional_formatting.add(
    f"G2:G{sheet.max_row}",
    ColorScaleRule(
        start_type="min",
        start_color="63BE7B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="F8696B",
    ),
)

green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
red = PatternFill("solid", fgColor="FFC7CE")

for row in range(2, sheet.max_row + 1):
    risk_cell = sheet[f"H{row}"]

    if risk_cell.value == "High":
        risk_cell.fill = red
    elif risk_cell.value == "Medium":
        risk_cell.fill = yellow
    elif risk_cell.value == "Low":
        risk_cell.fill = green

for cell in sheet["C"][1:]:
    cell.number_format = "$#,##0.00"

for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter

    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))

    sheet.column_dimensions[column_letter].width = max_length + 4

# Dashboard
if "Dashboard" in workbook.sheetnames:
    del workbook["Dashboard"]

dashboard = workbook.create_sheet("Dashboard", 0)
dashboard.sheet_view.showGridLines = False

dashboard.merge_cells("A1:F1")
dashboard["A1"] = "FinGuard Fraud Analytics Dashboard"
dashboard["A1"].font = Font(bold=True, size=24, color="1F4E78")
dashboard["A1"].alignment = Alignment(horizontal="center")

high_risk_count = len(transactions[transactions["Risk Level"] == "High"])
medium_risk_count = len(transactions[transactions["Risk Level"] == "Medium"])
low_risk_count = len(transactions[transactions["Risk Level"] == "Low"])
fraud_rate = high_risk_count / total_transactions

kpis = [
    ("Total Transactions", total_transactions),
    ("High Risk Transactions", high_risk_count),
    ("Medium Risk Transactions", medium_risk_count),
    ("Low Risk Transactions", low_risk_count),
    ("Fraud Rate", fraud_rate),
    ("Total Transaction Amount", total_amount),
    ("Average Transaction Amount", average_amount),
]

row = 3
for label, value in kpis:
    dashboard[f"A{row}"] = label
    dashboard[f"B{row}"] = value
    dashboard[f"A{row}"].font = Font(bold=True)
    row += 1

dashboard["B7"].number_format = "0.00%"
dashboard["B8"].number_format = "$#,##0.00"
dashboard["B9"].number_format = "$#,##0.00"

dashboard["A11"] = "Report Generated:"
dashboard["B11"] = datetime.now().strftime("%B %d, %Y %I:%M %p")
dashboard["A11"].font = Font(bold=True)

# Dashboard summary tables
risk_summary = transactions["Risk Level"].value_counts().reset_index()
risk_summary.columns = ["Risk Level", "Count"]

country_summary = transactions["Country"].value_counts().reset_index()
country_summary.columns = ["Country", "Count"]

payment_summary = transactions["PaymentMethod"].value_counts().reset_index()
payment_summary.columns = ["Payment Method", "Count"]

merchant_summary = transactions["Merchant"].value_counts().head(10).reset_index()
merchant_summary.columns = ["Merchant", "Count"]

category_summary = transactions["Category"].value_counts().reset_index()
category_summary.columns = ["Category", "Count"]


def write_table(start_col, title1, title2, data):
    dashboard.cell(row=3, column=start_col).value = title1
    dashboard.cell(row=3, column=start_col + 1).value = title2

    dashboard.cell(row=3, column=start_col).font = Font(bold=True)
    dashboard.cell(row=3, column=start_col + 1).font = Font(bold=True)

    for index, item in data.iterrows():
        dashboard.cell(row=index + 4, column=start_col).value = item[title1]
        dashboard.cell(row=index + 4, column=start_col + 1).value = item[title2]


write_table(4, "Risk Level", "Count", risk_summary)
write_table(7, "Country", "Count", country_summary)
write_table(10, "Payment Method", "Count", payment_summary)
write_table(13, "Merchant", "Count", merchant_summary)
write_table(16, "Category", "Count", category_summary)

charts = [
    ("charts/risk_levels.png", "A12"),
    ("charts/countries.png", "A34"),
    ("charts/payment_methods.png", "A56"),
    ("charts/top_merchants.png", "A78"),
    ("charts/transactions_by_category.png", "A100"),
]

for chart_path, location in charts:
    img = Image(chart_path)
    img.width = 720
    img.height = 420
    dashboard.add_image(img, location)

dashboard.column_dimensions["A"].width = 34
dashboard.column_dimensions["B"].width = 22
dashboard.column_dimensions["D"].width = 18
dashboard.column_dimensions["E"].width = 12
dashboard.column_dimensions["G"].width = 18
dashboard.column_dimensions["H"].width = 12
dashboard.column_dimensions["J"].width = 22
dashboard.column_dimensions["K"].width = 12
dashboard.column_dimensions["M"].width = 24
dashboard.column_dimensions["N"].width = 12
dashboard.column_dimensions["P"].width = 20
dashboard.column_dimensions["Q"].width = 12

workbook.save(output_file)

print("\n[SUCCESS] Professional Excel report saved!")